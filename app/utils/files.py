import hashlib
import os
import random
import string

import openpyxl
from fastapi import UploadFile

from app.core.config import settings


def create_dirs() -> None:
    for path in (
        settings.DATA_DIR,
        settings.IMAGES_DIR,
        settings.FILES_DIR,
        settings.IMPORTS_DIR,
        settings.ATTACHMENTS_DIR,
    ):
        os.makedirs(path, exist_ok=True)


def generate_random_filename(length: int = 10) -> str:
    characters = string.ascii_letters + string.digits
    return "".join(random.choice(characters) for _ in range(length))


def get_extension(filename: str | None) -> str:
    if not filename or "." not in filename:
        return ""
    return filename.rsplit(".", 1)[-1].lower()


async def read_upload_bytes(upload: UploadFile) -> bytes:
    content = await upload.read()
    await upload.seek(0)
    return content


async def hash_upload_file(upload: UploadFile) -> str:
    return hashlib.md5(await read_upload_bytes(upload)).hexdigest()


async def save_upload_file(upload: UploadFile, destination: str) -> None:
    content = await read_upload_bytes(upload)
    with open(destination, "wb") as file:
        file.write(content)


async def extract_text_from_upload(upload: UploadFile) -> str:
    content = await read_upload_bytes(upload)
    extension = get_extension(upload.filename)
    if extension in {"txt", "md", "csv", "json", "log", "xml", "html"}:
        return content.decode("utf-8", errors="ignore")
    if extension == "xlsx":
        workbook = openpyxl.load_workbook(filename=upload.file)
        try:
            parts: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    parts.append(" ".join(str(value) for value in row if value is not None))
            return "\n".join(filter(None, parts))
        finally:
            workbook.close()
            await upload.seek(0)
    return content.decode("utf-8", errors="ignore")


def extract_text_from_path(path: str | None) -> str:
    if not path or not os.path.exists(path):
        return ""
    extension = get_extension(path)
    if extension == "xlsx":
        workbook = openpyxl.load_workbook(filename=path)
        try:
            parts: list[str] = []
            for sheet in workbook.worksheets:
                for row in sheet.iter_rows(values_only=True):
                    parts.append(" ".join(str(value) for value in row if value is not None))
            return "\n".join(filter(None, parts))
        finally:
            workbook.close()
    with open(path, "rb") as file:
        return file.read().decode("utf-8", errors="ignore")


def delete_file_if_exists(path: str | None) -> None:
    if path and os.path.exists(path):
        os.remove(path)


async def load_image(upload: UploadFile, filename: str) -> None:
    await save_upload_file(upload, os.path.join(settings.IMAGES_DIR, filename))


async def delete_image(filename: str | None) -> None:
    delete_file_if_exists(os.path.join(settings.IMAGES_DIR, filename) if filename else None)
