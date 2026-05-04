from datetime import datetime

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment

from app.core.config import settings
from app.models.schemas import TodoExportRow
from app.models.schemas import TodoSource


def export_todos(todos: list, include_id: bool = True) -> str:
    workbook = Workbook()
    workbook.remove(workbook.active)
    worksheet = workbook.create_sheet("todos", 0)

    headers = [
        "title", "details", "completed", "tag",
        "created_at", "completed_at", "source",
        "image_path", "image_hash", "attachment_path",
    ]
    if include_id:
        headers.append("id")

    for index, header in enumerate(headers):
        worksheet.column_dimensions[chr(index + 65)].width = len(header) + 5
    worksheet.append(headers)
    for cell in worksheet[1]:
        cell.alignment = Alignment(horizontal="center")

    for todo in todos:
        tag_str = ", ".join(tag.name for tag in todo.tags) if hasattr(todo, "tags") else ""
        row = [
            todo.title,
            todo.details,
            "Выполнено" if todo.completed else "Не выполнено",
            tag_str,
            todo.created_at.strftime("%Y-%m-%d %H:%M:%S") if todo.created_at else "",
            todo.completed_at.strftime("%Y-%m-%d %H:%M:%S") if todo.completed_at else "",
            todo.source,
            todo.image_path,
            todo.image_hash,
            todo.attachment_path,
        ]
        if include_id:
            row.append(todo.id)
        worksheet.append(row)

    output_path = f"{settings.DATA_DIR}/todos.xlsx"
    workbook.save(output_path)
    return output_path


def import_todos(file_path: str) -> list[TodoExportRow]:
    workbook = openpyxl.load_workbook(file_path)
    sheet = workbook.active
    todos: list[TodoExportRow] = []

    for row in sheet.iter_rows(min_row=2, values_only=True):
        values = list(row)
        if len(values) == 10:
            title, details, completed, tag, created_at, completed_at, source, image_path, image_hash, attachment_path = values
            todo_id = None
        else:
            title, details, completed, tag, created_at, completed_at, source, image_path, image_hash, attachment_path, todo_id = values

        todos.append(
            TodoExportRow(
                id=todo_id,
                title=title,
                details=details or "",
                completed=completed == "Выполнено",
                tag=tag or "",
                created_at=created_at if isinstance(created_at, datetime) else None,
                completed_at=completed_at if isinstance(completed_at, datetime) else None,
                source=TodoSource.imported.value,
                image_path=image_path,
                image_hash=image_hash,
                attachment_path=attachment_path,
            )
        )

    workbook.close()
    return todos

